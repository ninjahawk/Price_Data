import os
import yfinance as yf
from openpyxl import load_workbook
from datetime import datetime

# Define Excel file path
file_path = 'Price Data Sheet_TEST.xlsx'

# Load workbook and access the "DROP DATA HERE" sheet
wb = load_workbook(file_path)
sheet = wb['DROP DATA HERE']

# Function to format today's date as required
def get_formatted_date():
    today = datetime.now()
    formatted_date = today.strftime("%a, %B %d, %Y")  # Mon, December 9, 2024
    return f"End-of-Day Recap - Price quotes for {formatted_date}"

# Update B3 with the formatted date
sheet['B3'] = get_formatted_date()

# Column Mapping
columns = {
    'Last': 3,     # Column C
    'Change': 4,   # Column D
    '%Chg': 5,     # Column E
    'Open': 6,     # Column F
    'High': 7,     # Column G
    'Low': 8,      # Column H
    'Volume': 9,   # Column I
    'Time': 10     # Column J
}

# Start processing from row 5 (data rows)
start_row = 5
end_row = sheet.max_row

def fetch_data(ticker):
    """Fetch data for the ticker."""
    try:
        stock = yf.Ticker(ticker)
        hist = stock.history(period="1d")  # Fetch today's data only
        if not hist.empty:
            return hist.iloc[-1]  # Return today's data
    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")
    return None

for row in range(start_row, end_row + 1):
    ticker = sheet.cell(row=row, column=2).value  # Column B - Ticker Symbol
    
    if ticker:
        print(f"Fetching data for {ticker}...")
        data = fetch_data(ticker)
        if data is not None:
            last_price = round(data['Close'], 2)
            open_price = round(data['Open'], 2)
            high_price = round(data['High'], 2)
            low_price = round(data['Low'], 2)
            volume = round(data['Volume'], 2)
            
            # Calculate Change and % Change based on Open price
            change = round(last_price - open_price, 2)
            percent_change = round((change / open_price) * 100, 2)

            # Write data to the correct columns
            sheet.cell(row=row, column=columns['Last'], value=last_price)
            sheet.cell(row=row, column=columns['Change'], value=change)
            sheet.cell(row=row, column=columns['%Chg'], value=percent_change)
            sheet.cell(row=row, column=columns['Open'], value=open_price)
            sheet.cell(row=row, column=columns['High'], value=high_price)
            sheet.cell(row=row, column=columns['Low'], value=low_price)
            sheet.cell(row=row, column=columns['Volume'], value=volume)
            sheet.cell(row=row, column=columns['Time'], value=datetime.now().strftime("%Y-%m-%d"))
        else:
            print(f"No sufficient data for {ticker}")
            sheet.cell(row=row, column=columns['Time'], value="Failed")

# Save the workbook back to the same file
try:
    wb.save(file_path)
    print(f"Data successfully updated and saved back to '{file_path}'.")
except Exception as e:
    print(f"Error saving workbook: {e}")
